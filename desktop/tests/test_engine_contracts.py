"""Engine edge-case + signature-contract tests (AGENT D).

These call the engine functions DIRECTLY — no Streamlit, no UI. They guard the
exact bug class this whole task targets: a drift between what the UI calls /
unpacks and what the engine actually accepts / returns.

desktop_app.py unpacks the engine returns positionally:
    data, nf, npairs, ndup = run_sor_xlsx_bytes(stage, title)   # 4-tuple
    data, nf, npairs       = run_trc_xlsx_bytes(stage, "...")    # 3-tuple
    data, nf, npairs       = run_json_xlsx_bytes(stage, "...")   # 3-tuple
(see desktop_app.py:272,274,295,297,308,310). If an engine refactor renamed
`folder`/`title` or changed a 4-tuple to a 3-tuple (or vice-versa), the UI would
crash at runtime with a ValueError nobody catches until a tech runs a job. The
signature tests below fail in <1 ms instead.

For every audit-flagged HIGH bug we pin BOTH:
  * a plain test that PASSES on today's (buggy) behaviour, tagged TODO, so the
    suite documents reality and won't silently rot; AND
  * an xfail(strict=True) that encodes the DESIRED behaviour, so the day the bug
    is fixed the xfail XPASSes and FAILS the build, forcing whoever fixes the
    engine to update this file in the same change.

Run from desktop/:   python -m pytest tests/test_engine_contracts.py
"""
import inspect
import os
import shutil
import tempfile

import pytest

import report
import report_sor
from conftest import FIXTURE_SOR_DIR, FIXTURE_JSON_DIR


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _param_names(func):
    return list(inspect.signature(func).parameters)


@pytest.fixture
def empty_dir():
    d = tempfile.mkdtemp()
    yield d
    shutil.rmtree(d, ignore_errors=True)


@pytest.fixture
def one_sor_dir():
    """A folder containing exactly ONE real .sor (below the >=2 minimum)."""
    d = tempfile.mkdtemp()
    src = os.path.join(FIXTURE_SOR_DIR, "TUCROM453_1550.sor")
    shutil.copy(src, os.path.join(d, "TUCROM453_1550.sor"))
    yield d
    shutil.rmtree(d, ignore_errors=True)


@pytest.fixture
def mixed_dir():
    """One folder holding BOTH 2 .sor and 2 .json — exercises per-engine glob
    isolation (each engine must see only its own extension)."""
    d = tempfile.mkdtemp()
    for fn in ("TUCROM453_1550.sor", "TUCROM454_1550.sor"):
        shutil.copy(os.path.join(FIXTURE_SOR_DIR, fn), os.path.join(d, fn))
    for fn in ("ELMMIL0001_1550 .json", "ELMMIL0002_1550 .json"):
        shutil.copy(os.path.join(FIXTURE_JSON_DIR, fn), os.path.join(d, fn))
    yield d
    shutil.rmtree(d, ignore_errors=True)


# ===========================================================================
# 1. SIGNATURE CONTRACTS — arg names the UI calls with
# ===========================================================================
# desktop_app.py calls every entry point POSITIONALLY as fn(stage, title), so
# the first two parameters MUST be (folder, title) in that order. A rename to
# e.g. (path, name) wouldn't break positional calls, but the kwargs the cloud
# twin / tests use would — and the names are the documented contract in MEMORY.

@pytest.mark.parametrize("func", [
    report_sor.run_sor_xlsx_bytes,
    report_sor.run_sor_bytes,
    report.run_json_xlsx_bytes,
    report.run_json_bytes,
    report.run_trc_xlsx_bytes,
    report.run_trc_bytes,
])
def test_entry_point_first_two_args_are_folder_then_title(func):
    names = _param_names(func)
    assert names[:2] == ["folder", "title"], (
        f"{func.__module__}.{func.__name__} first two params are {names[:2]!r}; "
        f"the UI calls it positionally as fn(stage, title) — a rename here is a "
        f"silent runtime break."
    )


def test_run_sor_entry_points_take_no_required_extra_args():
    # The UI calls run_sor_* with exactly (folder, title) and nothing else.
    # If a new REQUIRED param were added, those calls would TypeError.
    for func in (report_sor.run_sor_xlsx_bytes, report_sor.run_sor_bytes):
        params = inspect.signature(func).parameters.values()
        required = [p for p in params
                    if p.default is inspect.Parameter.empty
                    and p.kind in (p.POSITIONAL_ONLY, p.POSITIONAL_OR_KEYWORD)]
        assert [p.name for p in required] == ["folder", "title"], (
            f"{func.__name__} grew a new required arg the UI doesn't pass"
        )


# ===========================================================================
# 2. RETURN ARITY CONTRACTS — SOR=4-tuple, JSON/TRC=3-tuple
# ===========================================================================
# This is THE bug class. The UI does:
#   data, nf, npairs, ndup = run_sor_xlsx_bytes(...)   <- needs 4
#   data, nf, npairs       = run_json_xlsx_bytes(...)  <- needs 3
# A drift either way raises an uncaught ValueError mid-job.

def test_run_sor_xlsx_bytes_returns_4_tuple():
    out = report_sor.run_sor_xlsx_bytes(FIXTURE_SOR_DIR, "contract")
    assert isinstance(out, tuple) and len(out) == 4, (
        f"run_sor_xlsx_bytes returned {len(out) if isinstance(out, tuple) else type(out)}; "
        f"UI unpacks (data, nf, npairs, ndup)."
    )
    data, nf, npairs, ndup = out
    assert isinstance(data, (bytes, bytearray)) and len(data) > 0
    assert nf == 8 and npairs == 28          # 8 fixtures -> C(8,2)=28 pairs
    assert ndup == 1                         # TUCROM453~TUCROM454 is the 1 dup


def test_run_json_xlsx_bytes_returns_3_tuple():
    out = report.run_json_xlsx_bytes(FIXTURE_JSON_DIR, "contract")
    assert isinstance(out, tuple) and len(out) == 3, (
        f"run_json_xlsx_bytes returned {len(out) if isinstance(out, tuple) else type(out)}; "
        f"UI unpacks (data, nf, npairs)."
    )
    data, nf, npairs = out
    assert isinstance(data, (bytes, bytearray)) and len(data) > 0
    assert nf == 8 and npairs == 28


def test_sor_returns_one_more_element_than_json():
    """Cross-check the asymmetry directly: SOR must hand back exactly one extra
    field (the dup count) vs the JSON/TRC entry points the UI treats as 3-tuples.
    This catches the case where BOTH drift together to the same wrong arity."""
    s = report_sor.run_sor_xlsx_bytes(FIXTURE_SOR_DIR, "x")
    j = report.run_json_xlsx_bytes(FIXTURE_JSON_DIR, "x")
    assert len(s) == len(j) + 1 == 4


# ===========================================================================
# 3. _extract_fiber_num CONTRACT (small subset — NOT the full sweep)
# ===========================================================================
# The exhaustive 38k-pattern sweep lives in desktop/test_filenames.py and runs
# as its own CI step. Here we only pin the CONTRACT SEMANTICS that the engine
# relies on internally: canonical SS data parses, and an unparseable name
# returns None (sentinel) rather than 0 or raising — callers branch on None.

@pytest.mark.parametrize("fn,expected", [
    ("TUCROM453_1550.sor", 453),       # canonical Beta SOR
    ("ELMMIL0001_1550 .json", 1),      # canonical ELMMIL JSON (trailing space)
    ("LAGDUR.sor", None),              # no digits -> sentinel None
    ("Norsea_1550.sor", None),         # wavelength-only -> sentinel None
])
def test_extract_fiber_num_contract(fn, expected):
    assert report._extract_fiber_num(fn) == expected


def test_extract_fiber_num_unparseable_is_none_not_zero():
    # Contract: callers distinguish "no fiber number" (None sentinel) from a
    # parsed fiber 0. The sentinel must be None, never the int 0.
    got = report._extract_fiber_num("LAGDUR.sor")
    assert got is None
    assert not isinstance(got, int)


# ===========================================================================
# 4. EMPTY / SUB-MINIMUM FOLDERS — clear error, not a crash
# ===========================================================================
# Engine requires >=2 usable files. The contract is: raise a RuntimeError whose
# message names the shortfall, so the UI's `except Exception as e: st.error(e)`
# shows something a tech can act on — never a bare IndexError / segfault.

def test_empty_folder_sor_raises_clear_runtime_error(empty_dir):
    with pytest.raises(RuntimeError) as ei:
        report_sor.run_sor_xlsx_bytes(empty_dir, "x")
    assert "enough" in str(ei.value).lower() or "sor" in str(ei.value).lower()


def test_empty_folder_json_raises_clear_runtime_error(empty_dir):
    with pytest.raises(RuntimeError) as ei:
        report.run_json_xlsx_bytes(empty_dir, "x")
    assert "no json files" in str(ei.value).lower()


def test_empty_folder_trc_raises_clear_runtime_error(empty_dir):
    with pytest.raises(RuntimeError) as ei:
        report.run_trc_xlsx_bytes(empty_dir, "x")
    assert "no trc files" in str(ei.value).lower()


def test_single_sor_file_below_minimum_raises(one_sor_dir):
    with pytest.raises(RuntimeError) as ei:
        report_sor.run_sor_xlsx_bytes(one_sor_dir, "x")
    msg = str(ei.value).lower()
    assert "enough" in msg or "2" in msg, (
        f"one-file folder must explain the >=2 minimum, got: {ei.value!r}"
    )


# ===========================================================================
# 5. MIXED FORMATS IN ONE FOLDER — per-engine glob isolation
# ===========================================================================
def test_mixed_folder_json_engine_sees_only_json(mixed_dir):
    _, nf, _ = report.run_json_xlsx_bytes(mixed_dir, "x")
    assert nf == 2, f"JSON engine picked up non-json files: nf={nf}"


def test_mixed_folder_sor_engine_sees_only_sor(mixed_dir):
    _, nf, _, _ = report_sor.run_sor_xlsx_bytes(mixed_dir, "x")
    assert nf == 2, f"SOR engine picked up non-sor files: nf={nf}"


# ===========================================================================
# 6. AUDIT BUG — JSON XLSX hardcodes the 3-λ WL_ORDER label / columns
# ===========================================================================
# build_xlsx_json passes the module-global WL_ORDER=[1310,1550,1625] straight to
# build_xlsx_multiwl regardless of what the files carry (report.py:1972-1973).
# On a PURE-1550 folder the "Summary" sheet's Wavelengths cell reads
# "1310 nm, 1550 nm, 1625 nm" and the per-file sheet grows empty 1310/1625 span-
# loss columns. The fix (audit quick-win #2) is to intersect with the files' own
# wavelength list, mirroring the TRC path. Note the SAME workbook's "Acquisition
# parameters" sheet ALREADY reports the correct "All match: 1550" from the data,
# so the two sheets disagree — proof the Summary label is hardcoded, not derived.

import io
import openpyxl


def _json_xlsx_workbook():
    xb, _, _ = report.run_json_xlsx_bytes(FIXTURE_JSON_DIR, "wl")
    return openpyxl.load_workbook(io.BytesIO(xb))


def _summary_wavelengths_cell(wb):
    ws = wb["Summary"]
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "Wavelengths":
            return row[1]
    return None


def test_json_xlsx_summary_wavelengths_label_is_currently_hardcoded():
    """TODO(audit quick-win #2): pins the BUGGY label. On a pure-1550 folder the
    Summary sheet wrongly advertises all three wavelengths. Remove this and let
    the xfail below become the live assertion when build_xlsx_json starts
    intersecting WL_ORDER with the files' wavelengths."""
    wb = _json_xlsx_workbook()
    label = _summary_wavelengths_cell(wb)
    assert label == "1310 nm, 1550 nm, 1625 nm", (
        f"buggy hardcoded label changed to {label!r} — if this is the fix, "
        f"update/remove this test and flip the xfail."
    )
    # And prove the disagreement: the data-derived Acq sheet knows it's 1550.
    acq = wb["Acquisition parameters"]
    acq_vals = [c.value for r in acq.iter_rows() for c in r if c.value]
    assert any("All match: 1550" in str(v) for v in acq_vals)


@pytest.mark.xfail(strict=True, reason=(
    "AUDIT quick-win #2: build_xlsx_json hardcodes WL_ORDER=[1310,1550,1625] "
    "instead of intersecting with the files' wavelengths. On a pure-1550 folder "
    "the Summary 'Wavelengths' cell should read '1550 nm', not list 1310/1625."))
def test_json_xlsx_summary_wavelengths_should_match_data():
    wb = _json_xlsx_workbook()
    label = _summary_wavelengths_cell(wb)
    assert label == "1550 nm", (
        f"Summary should reflect only the wavelengths present (1550), got {label!r}"
    )


@pytest.mark.xfail(strict=True, reason=(
    "AUDIT quick-win #2: per-file verdict sheet should not emit span-loss "
    "columns for absent wavelengths. A pure-1550 folder should have no "
    "'Span loss @ 1310 nm' / '@ 1625 nm' header."))
def test_json_xlsx_per_file_has_no_absent_wavelength_columns():
    wb = _json_xlsx_workbook()
    ws = wb["Per-file verdict"]
    headers = [c.value for c in next(ws.iter_rows())]
    assert not any("1310" in str(h) for h in headers)
    assert not any("1625" in str(h) for h in headers)


# ===========================================================================
# 7. AUDIT HIGH #5 — name-key collapse silently drops distinct traces
# ===========================================================================
# load_file derives the dedup key as basename.split('_')[0] (report.py:124), so
# two DISTINCT multi-cable traces that share a prefix —
#   FTHNTXAD01_AD04_1550.json  and  FTHNTXAD01_AD05_1550.json
# both map to name 'FTHNTXAD01'. _warn_name_collisions then SKIPS the second,
# collapsing two real fibers into one with only a log-only WARN. The audit fix
# (cross-cutting #3) is a report.file_key() SSOT that keeps enough of the name to
# tell AD04 from AD05.

def _collide_pair():
    return [
        {"name": "FTHNTXAD01", "filepath": "/x/FTHNTXAD01_AD04_1550.json"},
        {"name": "FTHNTXAD01", "filepath": "/x/FTHNTXAD01_AD05_1550.json"},
        {"name": "OTHER",      "filepath": "/x/OTHER_1550.json"},
    ]


def test_name_key_currently_collapses_multicable_prefix():
    """TODO(audit HIGH#5): pins the BUGGY collapse. Two distinct AD04/AD05
    traces share name 'FTHNTXAD01' and the second is dropped — 3 in, 2 out."""
    kept = report._warn_name_collisions(_collide_pair())
    assert len(kept) == 2, (
        "expected the buggy collapse (one of AD04/AD05 dropped); if this changed "
        "the dedup key was fixed — flip the xfail below."
    )


def test_name_key_root_cause_is_split_underscore_not_the_extractor():
    """Root cause, pinned. The fiber-number extractor CAN tell AD04 from AD05
    (it returns 4 vs 5), so the data needed to disambiguate exists. The collapse
    comes purely from load_file deriving the dedup key as
    `basename.split('_')[0]` (report.py:124), which throws that distinction away
    and yields the SAME 'FTHNTXAD01' for both. Documents WHY the name key
    collapses and where the fix belongs."""
    # Extractor distinguishes them — not the culprit.
    assert (report._extract_fiber_num("FTHNTXAD01_AD04_1550.json")
            != report._extract_fiber_num("FTHNTXAD01_AD05_1550.json"))
    # But the load_file key derivation collapses them.
    key = lambda fn: os.path.basename(fn).split("_")[0].strip()
    assert key("FTHNTXAD01_AD04_1550.json") == key("FTHNTXAD01_AD05_1550.json") \
        == "FTHNTXAD01"


@pytest.mark.xfail(strict=True, reason=(
    "AUDIT HIGH#5 / cross-cutting #3: distinct multi-cable traces that share a "
    "prefix (FTHNTXAD01_AD04 vs _AD05) must NOT collapse to one dedup key. A "
    "report.file_key() SSOT should keep them distinct so neither is silently "
    "dropped."))
def test_multicable_traces_get_distinct_keys():
    # The audit fix is a public report.file_key() SSOT that keeps enough of the
    # name to tell AD04 from AD05. Today report.file_key does not exist, so this
    # raises AttributeError -> xfail. When the helper lands it must return
    # DISTINCT keys for the two traces (which is what stops the collapse), and
    # this xfail XPASSes -> FAILS the build, forcing a paired update here.
    assert hasattr(report, "file_key"), "report.file_key SSOT not implemented yet"
    k4 = report.file_key("/x/FTHNTXAD01_AD04_1550.json")
    k5 = report.file_key("/x/FTHNTXAD01_AD05_1550.json")
    assert k4 != k5, (
        f"file_key still collapses multi-cable traces: {k4!r} == {k5!r}"
    )
