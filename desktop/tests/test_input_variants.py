"""Input-variant tests for the Secret Sauce desktop app (AGENT C).

These exercise the *input plumbing* in front of the engines — the part that
turns whatever a tech drops in a folder (a flat folder, a .zip, a Mac-built
zip with AppleDouble junk, a Windows-built zip with backslash paths, a stray
non-trace .json, same-named files in sub-folders, the same fiber twice) into the
flat directory the engines glob. The contract under test is:

    folder of N real traces  ==  a .zip of those same N traces
                             ==  same engine verdict / same #confirmed-dups

plus the guards around it (AppleDouble skip, content-sniff, basename
de-collision, id-collision dedup). Where a path is *unguarded* it is pinned with
an xfail(strict=True) so the day it's hardened the test XPASSes and fails the
build, forcing a paired update.

Everything is driven through the SAME production flow a tech uses — streamlit's
AppTest with the no-op `__test_force_run__` hook — and through the importable
engine helpers (report._otdr_paths / report._warn_name_collisions,
report_sor._analyze_sor). Zip fixtures are built on the fly in tempfiles and
torn down; no .zip is committed.

The folder is injected by MUTATING THE WIDGET KEY ("folder") the path
text_input is bound to — Streamlit honours widget-key state over a side slot on
the next rerun, which is the reliable way to seed a value from a test.
"""
import io
import os
import glob
import shutil
import zipfile

import pytest
from openpyxl import load_workbook

import report
import report_sor
from conftest import run_streamlit, FIXTURE_SOR_DIR, FIXTURE_JSON_DIR


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------
def _sor_fixture_files():
    return sorted(glob.glob(os.path.join(FIXTURE_SOR_DIR, "*.sor")))


def _json_fixture_files():
    return sorted(glob.glob(os.path.join(FIXTURE_JSON_DIR, "*.json")))


def _drive_run(folder, want_xlsx=True):
    """Run the whole app against `folder` with output forced via the test hook.

    Returns the finished AppTest. `__test_force_run__` makes the Run branch fire
    without a button press (no-op for a real user); the folder is seeded onto the
    widget key so the path text_input picks it up on the first rerun.
    """
    at = run_streamlit(default_timeout=180)
    at.session_state["folder"] = folder
    at.session_state["__test_force_run__"] = True
    at.run()
    # The app caches the zip-extract temp dir in session_state and only cleans it
    # when the zip set changes; a test that builds a throwaway zip should not leak
    # the extract dir, so drop it here. NB: AppTest's session_state proxy treats
    # attribute/.get() access as a *widget-key* lookup and raises KeyError, so we
    # use `in` + bracket access rather than .get().
    if "_zip_dir" in at.session_state:
        shutil.rmtree(at.session_state["_zip_dir"], ignore_errors=True)
    return at


def _inventory_line(at):
    """The 'Found **N SOR** · ... ' success banner the inventory step emits."""
    for s in at.success:
        if "Found" in (s.value or ""):
            return s.value
    return ""


def _assert_clean_run(at):
    """No uncaught exception, no 'Analysis failed', and the 'Done' banner fired."""
    assert len(at.exception) == 0, (
        "app raised: " + repr([e.value for e in at.exception])
    )
    errs = [e.value for e in at.error]
    assert not any("Analysis failed" in (m or "") for m in errs), (
        "app reported analysis failure: " + repr(errs)
    )
    successes = [s.value for s in at.success]
    assert any("Done" in (m or "") for m in successes), (
        "run block did not complete — success banners were: " + repr(successes)
    )


def _confirmed_dup_rows(folder):
    """Non-empty rows of the 'Confirmed duplicates' sheet of the written report.

    Row 0 is the header, so a real dup pair means len >= 2. Using the actual
    written .xlsx (rather than an in-process return value) means the assertion
    covers the whole folder->stage->engine->write path the tech relies on.
    """
    out_dir = os.path.join(folder, "SecretSauce_reports")
    assert os.path.isdir(out_dir), f"SecretSauce_reports/ not created in {folder}"
    xlsxs = [
        os.path.join(out_dir, f)
        for f in os.listdir(out_dir)
        if f.lower().endswith(".xlsx")
    ]
    assert xlsxs, f"no .xlsx report written to {out_dir} (found {os.listdir(out_dir)})"
    wb = load_workbook(xlsxs[0])
    ws = wb["Confirmed duplicates"]
    return [
        [c for c in row if c is not None]
        for row in ws.iter_rows(values_only=True)
        if any(c is not None for c in row)
    ]


def _build_zip(zip_path, members):
    """Write a .zip at `zip_path`. `members` is a list of (arcname, src_path)."""
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for arc, src in members:
            zf.write(src, arcname=arc)


# ==========================================================================
# 1) folder  ==  zip  (same SOR data, same #confirmed-dups)
# ==========================================================================
def test_folder_and_zip_give_identical_dup_count(tmp_path):
    """The Beta SOR set produces the SAME confirmed-dup table whether it arrives
    as a flat folder or as a single .zip the test builds. This is the core
    folder==zip equivalence: _extract_zips_in + _stage_flat must hand the engine
    a directory indistinguishable from the flat folder, so the known
    TUCROM453~454 dup shows up exactly once either way."""
    sor = _sor_fixture_files()

    # (a) flat folder
    folder_dir = tmp_path / "flat"
    folder_dir.mkdir()
    for f in sor:
        shutil.copy(f, str(folder_dir / os.path.basename(f)))
    at_folder = _drive_run(str(folder_dir))
    _assert_clean_run(at_folder)
    folder_rows = _confirmed_dup_rows(str(folder_dir))

    # (b) the very same bytes, delivered as one zip in an otherwise empty folder
    zip_holder = tmp_path / "zipped"
    zip_holder.mkdir()
    _build_zip(
        str(zip_holder / "romero.zip"),
        [(os.path.basename(f), f) for f in sor],
    )
    at_zip = _drive_run(str(zip_holder))
    _assert_clean_run(at_zip)
    zip_rows = _confirmed_dup_rows(str(zip_holder))

    # The zip inventory must announce it saw the archive ...
    assert "zip archive" in _inventory_line(at_zip), (
        "zip run did not report a zip archive: " + _inventory_line(at_zip)
    )
    assert "8 SOR" in _inventory_line(at_zip), _inventory_line(at_zip)

    # ... and the verdict must be byte-for-byte the same shape: same #rows, and
    # both naming the one real dup pair.
    assert len(folder_rows) == len(zip_rows) == 2, (
        f"folder gave {len(folder_rows)} rows, zip gave {len(zip_rows)} rows "
        "(expected header + the single TUCROM453~454 dup in both)"
    )
    for label, rows in (("folder", folder_rows), ("zip", zip_rows)):
        flat = " ".join(str(c) for r in rows[1:] for c in r)
        assert "TUCROM453" in flat and "TUCROM454" in flat, (
            f"{label} report does not list the known TUCROM453~454 dup: {rows[1:]}"
        )


# ==========================================================================
# 2) AppleDouble ._* / __MACOSX junk in a Mac-built zip is skipped
# ==========================================================================
def test_macstyle_zip_appledouble_entries_skipped(tmp_path):
    """A zip built on macOS carries a __MACOSX/._NAME AppleDouble sidecar next to
    every real file. Those are not traces and crash the parsers, so the input
    layer must drop them: _extract_zips_in skips inner names starting with '._'
    (and .DS_Store), and the recursive inventory + report._otdr_paths skip any
    '._' basename. End result: exactly the 8 real SOR, same dup verdict as a
    clean folder — no parser crash, no phantom files."""
    sor = _sor_fixture_files()
    holder = tmp_path / "mac"
    holder.mkdir()
    members = []
    for f in sor:
        b = os.path.basename(f)
        members.append((b, f))
    zip_path = str(holder / "mac_built.zip")
    # Build the real members first, then inject Mac junk entries by hand so the
    # AppleDouble sidecars sit *inside* the same zip alongside the traces.
    _build_zip(zip_path, members)
    with zipfile.ZipFile(zip_path, "a", zipfile.ZIP_DEFLATED) as zf:
        for f in sor:
            b = os.path.basename(f)
            zf.writestr("__MACOSX/._" + b, b"Mac AppleDouble metadata, not a trace")
        zf.writestr(".DS_Store", b"\x00\x00 mac finder junk \x00")

    at = _drive_run(str(holder))
    _assert_clean_run(at)
    assert "8 SOR" in _inventory_line(at), (
        "AppleDouble/__MACOSX junk leaked into the inventory: " + _inventory_line(at)
    )
    rows = _confirmed_dup_rows(str(holder))
    assert len(rows) == 2, (
        f"expected the clean 8-SOR verdict from a Mac zip, got {len(rows)} rows"
    )


def test_otdr_paths_skips_appledouble_directly(tmp_path):
    """Unit-level twin of the above: report._otdr_paths must drop '._*' sidecars
    even when they carry a real OTDR extension, so the engine never tries to
    parse one. Guards 05eabe2 directly."""
    d = tmp_path / "ad"
    d.mkdir()
    for f in _sor_fixture_files():
        shutil.copy(f, str(d / os.path.basename(f)))
    # AppleDouble sidecar with a .sor extension — a naive glob('*.sor') catches it.
    (d / "._TUCROM453_1550.sor").write_bytes(b"not a trace")
    paths = report._otdr_paths(str(d), "*.sor")
    assert len(paths) == 8, f"expected 8 real .sor, got {len(paths)}: {paths}"
    assert all(not os.path.basename(p).startswith("._") for p in paths), (
        "a '._' AppleDouble sidecar survived _otdr_paths: " + repr(paths)
    )


# ==========================================================================
# 3) Windows-built zip: backslash separators in entry names
# ==========================================================================
def test_windows_backslash_zip_entries_extract(tmp_path):
    """A zip produced by some Windows tools stores entry names with BACKSLASH
    separators, e.g. 'job1\\TUCROM449_1550.sor'. On macOS/Linux those are not
    treated as path separators, so os.path.basename keeps the whole 'job1\\...'
    string. _extract_zips_in takes os.path.basename(info.filename) for the dest
    name, so it must still recover the real trace files and run cleanly."""
    sor = _sor_fixture_files()
    holder = tmp_path / "win"
    holder.mkdir()
    zip_path = str(holder / "win_built.zip")
    # writestr with an explicit backslash arcname forces the Windows-style name
    # into the central directory regardless of host OS.
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in sor:
            zf.writestr("job1\\" + os.path.basename(f), open(f, "rb").read())

    # Sanity: the entries really do carry backslashes (so the test exercises the
    # quirk, not a silently-normalised name).
    with zipfile.ZipFile(zip_path) as zf:
        names = [i.filename for i in zf.infolist()]
    assert any("\\" in n for n in names), (
        "test setup failed: zip entries were normalised, no backslash present: "
        + repr(names)
    )

    at = _drive_run(str(holder))
    _assert_clean_run(at)
    assert "8 SOR" in _inventory_line(at), (
        "backslash-named zip entries were not recovered: " + _inventory_line(at)
    )
    rows = _confirmed_dup_rows(str(holder))
    assert len(rows) == 2, (
        f"expected the 8-SOR dup verdict from a Windows-built zip, got {len(rows)} "
        "rows"
    )


# ==========================================================================
# 4) content-sniff: a garbage .json is rejected, not run
# ==========================================================================
def test_content_sniff_rejects_non_trace_json(tmp_path):
    """A stray results/config .json (no 'OtdrMeasurements' key) sitting next to
    real traces must be ignored by the content-sniff (_is_otdr_json), NOT fed to
    the engine and NOT counted toward the one-type-per-run guard. The inventory
    should report the 8 real traces and surface an 'Ignored ... non-trace .json'
    note, and the run should complete."""
    folder = tmp_path / "json"
    folder.mkdir()
    for f in _json_fixture_files():
        shutil.copy(f, str(folder / os.path.basename(f)))
    # garbage json: valid JSON, but not an OTDR trace
    (folder / "rmse_results.json").write_text(
        '{"rmse": 0.42, "note": "results export, not a trace"}'
    )
    # outright non-JSON bytes with a .json extension — must also be rejected, and
    # must not crash the sniff (it reads bytes, no parse).
    (folder / "broken.json").write_bytes(b"\x00\x01 not even text \xff\xfe")

    at = _drive_run(str(folder))
    _assert_clean_run(at)

    line = _inventory_line(at)
    assert "8 JSON" in line, f"content-sniff dropped a real trace or kept junk: {line}"
    captions = [c.value for c in at.caption]
    assert any("Ignored" in (c or "") and "non-trace" in (c or "") for c in captions), (
        "no 'Ignored N non-trace .json' note surfaced; captions were: "
        + repr(captions)
    )


# ==========================================================================
# 5) A == B  /  same fiber twice  (self-duplicate)
# ==========================================================================
def test_identical_file_twice_is_a_confirmed_dup(tmp_path):
    """A==B: the exact same SOR bytes under two distinct filenames is, by
    definition, a 100% duplicate — there is no self-pair suppression, and there
    shouldn't be (identical content IS a re-shoot). _analyze_sor must score the
    single pair at p_dup==1.0 and count it (n50==1)."""
    d = tmp_path / "ab"
    d.mkdir()
    src = os.path.join(FIXTURE_SOR_DIR, "TUCROM453_1550.sor")
    shutil.copy(src, str(d / "COPYA_1550.sor"))
    shutil.copy(src, str(d / "COPYB_1550.sor"))

    res = report_sor._analyze_sor(str(d))
    files = res["files"]
    assert len(files) == 2, f"expected both copies to load, got {len(files)}"
    pdup = list(res["p_dup"])
    assert len(pdup) == 1 and pdup[0] == pytest.approx(1.0), (
        f"identical bytes must score p_dup==1.0, got {pdup}"
    )
    assert res["n50"] == 1, f"identical-file pair not counted as a dup: n50={res['n50']}"


# ==========================================================================
# 6) sub-directory basename collisions
#    SOR: full-stem key -> _stage_flat rename keeps BOTH distinct files.
#    JSON: split('_')[0] key -> distinct same-prefix files COLLAPSE via the
#          _warn_name_collisions guard (which emits a visible WARN).
# ==========================================================================
def test_sor_subdir_basename_collision_keeps_both(tmp_path):
    """Two DIFFERENT SOR traces that happen to share a basename, sitting in
    different sub-folders, must BOTH survive. The SOR engine keys off the full
    filename stem, and the desktop app's _stage_flat renames the second colliding
    basename to '<stem>__1.sor', so both reach the engine. We assert the engine
    sees 2 files (not 1) when fed the staged dir the app would build."""
    sub_a = tmp_path / "site_a"
    sub_b = tmp_path / "site_b"
    sub_a.mkdir()
    sub_b.mkdir()
    sor = _sor_fixture_files()
    # Same basename 'FIB001_1550.sor', different real content in each subdir.
    shutil.copy(sor[0], str(sub_a / "FIB001_1550.sor"))
    shutil.copy(sor[1], str(sub_b / "FIB001_1550.sor"))

    # Reproduce the app's _stage_flat basename de-collision inline (importing the
    # app module is not possible — it st.stop()s at module top — so we mirror the
    # documented rename: second colliding basename -> '<stem>__1<ext>').
    staged = tmp_path / "staged"
    staged.mkdir()
    used = set()
    for p in [str(sub_a / "FIB001_1550.sor"), str(sub_b / "FIB001_1550.sor")]:
        base = os.path.basename(p)
        dest = base
        i = 1
        while dest.lower() in used:
            stem, ext = os.path.splitext(base)
            dest = f"{stem}__{i}{ext}"
            i += 1
        used.add(dest.lower())
        shutil.copy(p, str(staged / dest))

    on_disk = report._otdr_paths(str(staged), "*.sor")
    assert len(on_disk) == 2, (
        f"_stage_flat de-collision should keep both SOR files, found {on_disk}"
    )
    res = report_sor._analyze_sor(str(staged))
    assert len(res["files"]) == 2, (
        "two distinct same-basename SOR traces collapsed to "
        f"{len(res['files'])}; the full-stem key + rename should keep both"
    )


def test_json_same_prefix_distinct_files_collapse_with_warn(tmp_path, capsys):
    """Counterpart for JSON/TRC, where the engine name-key is basename.split('_')[0].
    Two DISTINCT json traces whose names share that prefix map to the SAME id, so
    even after a basename rename _warn_name_collisions DROPS the second and keeps
    the first. This is the documented guard: we assert (a) it dedups (4 distinct
    files -> 3 kept) and (b) it emits a VISIBLE 'duplicate fiber id ... SKIPPING'
    WARN rather than silently losing the trace."""
    src = _json_fixture_files()
    d = tmp_path / "j"
    d.mkdir()
    # Two distinct traces, both forced to prefix 'ELMMIL0001' (-> same id key).
    shutil.copy(src[0], str(d / "ELMMIL0001_1550 .json"))
    shutil.copy(src[1], str(d / "ELMMIL0001_alt_1550 .json"))
    # Plus two genuinely distinct files so >=2 survive for a valid run.
    shutil.copy(src[2], str(d / os.path.basename(src[2])))
    shutil.copy(src[3], str(d / os.path.basename(src[3])))

    paths = report._otdr_paths(str(d), "*.json")
    assert len(paths) == 4, f"expected 4 files on disk, got {len(paths)}"
    loaded = []
    for p in paths:
        try:
            loaded.append(report.load_file(p))
        except Exception as exc:  # pragma: no cover - fixtures are valid traces
            pytest.fail(f"fixture {p} failed to load: {exc}")
    assert sorted(f["name"] for f in loaded).count("ELMMIL0001") == 2, (
        "test setup: the two files should share id 'ELMMIL0001'"
    )

    kept = report._warn_name_collisions(loaded)
    assert len(kept) == 3, (
        f"id-collision dedup should keep 3 of 4 distinct files, kept {len(kept)}"
    )
    # The drop must be VISIBLE, not silent.
    out = capsys.readouterr().out
    assert "duplicate fiber id" in out and "SKIPPING" in out, (
        "the dropped same-id trace was not surfaced via a WARN; stdout was:\n" + out
    )


# ==========================================================================
# 7) XFAIL pins: same-name files inside ONE zip overwrite (audit HIGH#2)
# ==========================================================================
@pytest.mark.xfail(
    strict=True,
    reason="Audit HIGH#2: _extract_zips_in writes inner files by basename into "
    "one per-zip subdir, so two same-named files in DIFFERENT folders inside the "
    "SAME zip overwrite each other before inventory sees them. When a within-zip "
    "de-collision counter lands, both will survive and this XPASSes.",
)
def test_same_name_within_one_zip_should_not_overwrite(tmp_path):
    """job1/FIB001_1550.sor and job2/FIB001_1550.sor — two DISTINCT traces, same
    basename, both inside ONE zip. _extract_zips_in makes one dest subdir per ZIP
    (not per inner folder) and writes by os.path.basename, so the second silently
    overwrites the first: the engine then sees ONE file where there were two.
    Pinned xfail(strict): the day a within-zip counter/suffix is added, both
    survive, the inventory shows 2 SOR (plus the 6 others), and this XPASSes,
    failing the build to force a paired update."""
    sor = _sor_fixture_files()
    holder = tmp_path / "onezip"
    holder.mkdir()
    zip_path = str(holder / "two_jobs.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        # Same basename, different real content, different inner folders.
        zf.writestr("job1/FIB001_1550.sor", open(sor[0], "rb").read())
        zf.writestr("job2/FIB001_1550.sor", open(sor[1], "rb").read())
        # 6 more genuinely distinct traces so a run is always valid either way.
        for f in sor[2:]:
            zf.writestr("job1/" + os.path.basename(f), open(f, "rb").read())

    at = _drive_run(str(holder))
    _assert_clean_run(at)
    line = _inventory_line(at)
    # If the overwrite is fixed there are 8 SOR; today the collision drops one -> 7.
    assert "8 SOR" in line, (
        "within-zip same-name overwrite dropped a trace; inventory: " + line
    )
