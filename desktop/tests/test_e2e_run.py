"""End-to-end "Run analysis" tests for the Secret Sauce desktop app.

These drive the WHOLE flow through streamlit's AppTest exactly as a tech would:
point the app at a folder of real OTDR files, trigger the run, and assert that a
real .xlsx report lands in the folder's SecretSauce_reports/ subfolder with the
sheets the engine is supposed to produce.

The run is triggered headlessly via the production test hook in desktop_app.py:
setting st.session_state["__test_force_run__"] = True makes `run_clicked` true
without a real button press. That flag is never set by any UI element, so this
is a no-op for a real user.

The folder is supplied by MUTATING THE WIDGET KEY ("folder") that the path
text_input is bound to — Streamlit honours widget-key state over a side slot on
the next rerun, so this is the reliable way to inject a value from a test.

Each test copies its committed fixture set into a fresh temp dir so the report
gets written to throwaway space and the committed fixtures are never polluted.
"""
import io
import os
import shutil

import pytest
from openpyxl import load_workbook

from conftest import run_streamlit, FIXTURE_SOR_DIR, FIXTURE_JSON_DIR

# Sheets the engine emits, in the order the desktop report is supposed to show
# them. "Acquisition parameters" is inserted at index 0 (first tab) by
# report._write_acq_sheet; the core tables follow.
EXPECTED_SHEETS_IN_ORDER = [
    "Acquisition parameters",
    "Summary",
    "Per-file verdict",
    "Confirmed duplicates",
]


def _copy_fixture(src_dir, tmp_path, name):
    """Copy a fixture folder into throwaway temp space and return the copy."""
    dest = os.path.join(str(tmp_path), name)
    shutil.copytree(src_dir, dest)
    return dest


def _drive_run(folder):
    """Run the app against `folder` with Excel output forced via the test hook.

    Returns the finished AppTest.
    """
    at = run_streamlit(default_timeout=120)
    # Seed the folder onto the WIDGET KEY before the first run so the path
    # text_input picks it up (widget-key state wins over the side slot).
    at.session_state["folder"] = folder
    # Force the Run branch without a real button click. No-op in production.
    at.session_state["__test_force_run__"] = True
    at.run()
    return at


def _assert_no_exception(at):
    """AppTest collects uncaught exceptions in at.exception (an ElementList,
    never None). Empty == clean run. Also confirm the run branch actually fired
    by checking for the app's "Done" success banner, so a silently-skipped run
    (e.g. an st.stop() before the report) can't masquerade as a pass."""
    assert len(at.exception) == 0, (
        "app raised: " + repr([e.value for e in at.exception])
    )
    # The app must NOT have surfaced an "Analysis failed: ..." error.
    errs = [e.value for e in at.error]
    assert not any("Analysis failed" in (m or "") for m in errs), (
        "app reported analysis failure: " + repr(errs)
    )
    # And it must have reached the success banner that the run block emits.
    successes = [s.value for s in at.success]
    assert any("Done" in (m or "") for m in successes), (
        "run block did not complete — success banners were: " + repr(successes)
    )


def _find_report_xlsx(folder):
    """Return the path to the produced report .xlsx under SecretSauce_reports/."""
    out_dir = os.path.join(folder, "SecretSauce_reports")
    assert os.path.isdir(out_dir), (
        f"SecretSauce_reports/ was not created in {folder}"
    )
    xlsxs = [
        os.path.join(out_dir, f)
        for f in os.listdir(out_dir)
        if f.lower().endswith(".xlsx")
    ]
    assert xlsxs, f"no .xlsx report written to {out_dir} (found {os.listdir(out_dir)})"
    return xlsxs[0]


def _assert_expected_sheets(xlsx_path):
    wb = load_workbook(xlsx_path)
    names = wb.sheetnames
    for sheet in EXPECTED_SHEETS_IN_ORDER:
        assert sheet in names, f"missing sheet {sheet!r}; got {names!r}"
    # "Acquisition parameters" must be the FIRST tab.
    assert names[0] == "Acquisition parameters", (
        f"first sheet should be 'Acquisition parameters', got {names[0]!r}"
    )
    # The core tables must appear in the documented relative order.
    positions = [names.index(s) for s in EXPECTED_SHEETS_IN_ORDER]
    assert positions == sorted(positions), (
        f"core sheets out of expected order: {[names[p] for p in positions]}"
    )
    return names


def test_e2e_run_sor(tmp_path):
    """Full flow on the real Beta ROMERO->TUCUMCARI SOR fixture (has 1 dup)."""
    folder = _copy_fixture(FIXTURE_SOR_DIR, tmp_path, "sor")
    at = _drive_run(folder)

    _assert_no_exception(at)

    xlsx_path = _find_report_xlsx(folder)
    _assert_expected_sheets(xlsx_path)

    # The TUCROM453 ~ TUCROM454 pair is a genuine ~100% duplicate, so the
    # "Confirmed duplicates" sheet must carry at least one detail row beyond its
    # header — and that row must name the real dup pair.
    wb = load_workbook(xlsx_path)
    ws = wb["Confirmed duplicates"]
    rows = [
        [c for c in row if c is not None]
        for row in ws.iter_rows(values_only=True)
        if any(c is not None for c in row)
    ]
    assert len(rows) >= 2, (
        "expected the real TUCROM453~454 dup to populate the Confirmed "
        f"duplicates sheet (header + >=1 detail row), but got {len(rows)} "
        "non-empty row(s)"
    )
    flat = " ".join(str(c) for r in rows[1:] for c in r)
    assert "TUCROM453" in flat and "TUCROM454" in flat, (
        "Confirmed duplicates sheet does not list the known TUCROM453~454 "
        f"dup pair; detail rows were: {rows[1:]}"
    )


def test_e2e_run_json(tmp_path):
    """Full flow on the real ELMMIL single-wavelength JSON fixture."""
    folder = _copy_fixture(FIXTURE_JSON_DIR, tmp_path, "json")
    at = _drive_run(folder)

    _assert_no_exception(at)

    xlsx_path = _find_report_xlsx(folder)
    _assert_expected_sheets(xlsx_path)
